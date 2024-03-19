import tkinter as tk
from tkinter import filedialog
from docx import Document
import pandas as pd
from openpyxl import load_workbook
import re

def procesar_archivo():
    import re
    # Abrir el cuadro de diálogo para seleccionar el archivo
    ruta_archivo = filedialog.askopenfilename()
    
    # Verificar si se seleccionó un archivo
    if ruta_archivo:

        # Ruta del archivo de texto original y del archivo de salida modificado
        ruta_archivo_txt_original=ruta_archivo
        ruta_archivo_txt_modificado = 'archivo_modificado.txt'
        ruta_archivo_excel_modificado = 'excel_final.xlsx'

        # Leer el contenido del archivo de texto original
        with open(ruta_archivo_txt_original, 'r', encoding='utf-8') as archivo_txt:
            contenido_txt_original = archivo_txt.read()

        # Eliminar las URLs del contenido del archivo de texto original
        contenido_txt_sin_urls = eliminar_urls(contenido_txt_original)

        # Eliminar todo el texto después de encontrar un asterisco (*) o dos asteriscos (**) en el archivo de texto
        # Eliminar todo el texto desde los asteriscos hasta "2024", incluyendo el salto de línea
        contenido_txt_sin_asteriscos = re.sub(r'\*{1,2}.*?2024.*?[\r\n]*', '', contenido_txt_sin_urls, flags=re.DOTALL)


        # Guardar el contenido modificado en un nuevo archivo de texto
        with open(ruta_archivo_txt_modificado, 'w', encoding='utf-8') as archivo_salida:
            archivo_salida.write(contenido_txt_sin_asteriscos)

        # 1. Eliminar texto de títulos y juntas
        buscar_reemplazar_titulos = [
            ("PROCESO ELECTORAL CONCURRENTE 2023-2024", ""),
            ("PROCESO ELECTORAL 2023-2024", " "),
            ("CÉDULA DE INFORMACIÓN CRyT ITINERANTE", ""),
            ("MAPA DE LA RUTA PREFERENTE", ""),
            ("MUNICIPIO:", "MUNICIPIO"),
            ("\nB)", "B)"),
            ("\n B)", "B)"),
            ("km.", "km"),
            ("Ayuntamientos: ", "Ayuntamientos"),
            (":", " "),
            ("Domicilio", ""),
            
            # Se pueden añadir más títulos y reemplazos si es necesario
        ]

        buscar_reemplazar_juntas = [
            ("JUNTA DISTRITAL EJECUTIVA 00 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 01 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 02 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 03 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 04 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 05 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 06 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 07 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 08 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 09 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 10 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 11 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 12 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 13 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 14 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 15 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 16 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 17 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 18 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 19 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 20 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 21 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 22 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 23 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 24 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 25 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 26 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 27 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 28 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 29 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 30 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 31 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 32 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 33 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 34 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 35 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 36 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 37 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 38 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 39 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 40 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 41 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 42 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 43 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 44 ESTADO DE MÉXICO", ""),
            ("JUNTA DISTRITAL EJECUTIVA 45 ESTADO DE MÉXICO", "")
            #VARIABLE
        ]

        with open(ruta_archivo_txt_modificado, 'r', encoding='utf-8') as archivo_txt:
            contenido_txt = archivo_txt.read()

        for buscar, reemplazar in buscar_reemplazar_titulos:
            contenido_txt = contenido_txt.replace(buscar, reemplazar)
        for buscar, reemplazar in buscar_reemplazar_juntas:
            contenido_txt = contenido_txt.replace(buscar, reemplazar)

        # 2. Eliminar texto de notas
        notas_a_eliminar = [
            "* Esta es una ruta sugerida que podría modificarse de acuerdo a la conclusión del escrutinio y cómputo de las casillas que la comprenden.",
            "* Esta es una ruta sugerida que podría modificarse de acuerdo con la conclusión del escrutinio y cómputo de las casillas que la comprenden.",
            "**Las Representaciones de los Partidos Políticos o Candidaturas Independientes podrán acompañar y vigilar, por sus propios medios, el recorrido del mecanismo de recolección hasta la entrega de los paquetes electorales a la sede del consejo correspondiente. (Artículo 334, párrafo 1, inciso e), Reglamento de Elecciones).",
            "** Las Representaciones de los Partidos Políticos o Candidaturas Independientes podrán acompañar y vigilar, por sus propios medios, el recorrido del mecanismo de recolección hasta la entrega de los paquetes electorales a la sede del consejo correspondiente. (Artículo 334, párrafo 1, inciso e), Reglamento de Elecciones)."]

        for nota in notas_a_eliminar:
            contenido_txt = contenido_txt.replace(nota, "")

        # 3. Reemplazar saltos de línea con espacio y eliminar dobles espacios
        contenido_txt = contenido_txt.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
        contenido_txt = re.sub(r' +', ' ', contenido_txt)

        # Guardar el contenido modificado en un nuevo archivo
        with open('archivo_modificado.txt', 'w', encoding='utf-8') as archivo_salida:
            archivo_salida.write(contenido_txt)

        ruta_archivo_txt = 'archivo_modificado.txt'

        # 5. Dividir cédulas
        # a) Reemplazar texto ZORE con salto de línea manual
        with open(ruta_archivo_txt, 'r', encoding='utf-8') as archivo_txt:
            contenido_txt = archivo_txt.read()

        contenido_txt = contenido_txt.replace("ZORE", "^IZORE")

        # 6. Agregar signo de separación al texto restante
        # Reemplazar espacios en blanco con el signo ">"
        contenido_txt = contenido_txt.replace(" ", ">")

        # Guardar el contenido modificado en un nuevo archivo
        with open('archivo_modificado_.txt', 'w', encoding='utf-8') as archivo_salida:
            archivo_salida.write(contenido_txt)

        import pandas as pd

        # Leer el contenido modificado del archivo de texto
        ruta_archivo_txt = 'archivo_modificado_.txt'

        with open(ruta_archivo_txt, 'r', encoding='utf-8') as archivo_txt:
            contenido_txt = archivo_txt.read()

        # Dividir el contenido por el delimitador ">"
        cedulas_divididas = contenido_txt.split('>')

        # Crear listas para columnas en DataFrame
        columna_1 = []
        columna_2 = []

        # Iterar sobre las cédulas divididas
        for cedula in cedulas_divididas:
            if 'ZOR' in cedula:
                # Agregar nueva fila en caso de encontrar "ZOR"
                columna_1.append(cedula)
                columna_2.append('')
            else:
                # Agregar cédula a la última fila
                if columna_2:  # Asegurarse de agregar espacio solo si ya hay elementos en columna_2
                    columna_2[-1] += ' ' + cedula
                else:
                    # Agregar cédula a una nueva fila si columna_2 está vacía
                    columna_2.append(cedula)

        # Asegurarse de que ambas columnas tengan la misma longitud
        if len(columna_1) < len(columna_2):
            columna_1.append('')

        # Crear un DataFrame de pandas con dos columnas
        df = pd.DataFrame({'Columna 1': columna_1, 'Columna 2': columna_2})

        # Guardar el DataFrame en un archivo Excel
        ruta_archivo_excel = 'archivo_modificado_.xlsx'
        df.to_excel(ruta_archivo_excel, index=False)

        # El archivo Excel se guarda, y ahora puedes abrirlo en Excel
        #*********************AGREGAR > ESPECÍFICOS*******************************************************

        import re
        from openpyxl import load_workbook

        # Textos específicos que deseas buscar
        textos_especificos = [
            "DISTRITO FEDERAL",
            "DISTRITO LOCAL",
            "MUNICIPIO",
            "TIPO Y NÚMERO DE MECANISMO",
            "TIPO DE ELECCIÓN",
            "NÚMERO Y TIPO DE CASILLAS",
            "TOTAL DE PAQUETES",
            "PUNTO DE PARTIDA PREFERENTE DEL RECORRIDO",
            "DESTINO(S) INMEDIATO(S)",
            "DESTINO FINAL ENTREGA DE PAQUETES",
            "NOMBRE DEL TITULAR Y TELÉFONO ÓRGANO ELECTORAL",
            "PERSONA RESPONSABLE Y/O AUXILIAR DEL MECANISMO",
            "DATOS DEL VEHÍCULO",
            "Local: Ayuntamiento",
            "Local: Ayuntamientos",
            "Local: Diputación",
            "No aplica",
            "PAQUETES POR TIPO DE ELECCIÓN"
        ]

        # Ruta del archivo Excel
        ruta_archivo_excel = 'archivo_modificado_.xlsx'

        # Cargar el libro de trabajo de Excel
        wb = load_workbook(filename=ruta_archivo_excel)
        # Seleccionar la primera hoja del libro de trabajo
        ws = wb.active


        for row_idx, row in enumerate(ws.iter_rows()):
            for cell in row:
                contenido_celda = str(cell.value)
                
                # Identificar la expresión 'km > >NOMBRE' en una celda de la fila
                if 'km > >NOMBRE' in contenido_celda:
                    # Insertar una fila vacía antes de la fila actual
                    ws.insert_rows(row_idx)
                    break  # Salir del bucle para evitar duplicaciones si hay múltiples celdas con 'km > >NOMBRE'
            
            # Recalcular el índice de la fila para evitar problemas con la inserción de filas
            row_idx += 1
            
        # Iterar sobre todas las celdas de la hoja de trabajo
        for row in ws.iter_rows():
            for cell in row:
                # Leer el contenido de la celda
                contenido_celda = str(cell.value)
                
                # Iterar sobre los textos específicos y agregar el delimitador ">"
                for texto in textos_especificos:
                    # Construir la expresión regular para buscar el texto específico con o sin espacios alrededor
                    regex = re.compile(rf'(?i)(\b{re.escape(texto)}\b)')
                    # Reemplazar el texto encontrado con el mismo texto precedido y seguido por ">"
                    contenido_celda = regex.sub(r'>\1>', contenido_celda)

                # Agregar ">" antes de "ARE" si hay un número antes de "ARE"
                contenido_celda = re.sub(r'(\b\d+\b)\s+ARE\s+(\d+\b)', r'\1 > ARE \2', contenido_celda)
        
        
                # Agregar ">" antes y después del número de paquetes
                contenido_celda = re.sub(r'(\d+)\s+(PAQUETES)', r'>\1 PAQUETES >', contenido_celda)

                # Buscar " PAQUETES POR >TIPO DE ELECCIÓN>" y reemplazar por " PAQUETES POR TIPO DE ELECCIÓN>"
                contenido_celda = contenido_celda.replace(" PAQUETES POR >TIPO DE ELECCIÓN>", " PAQUETES POR TIPO DE ELECCIÓN>")
        
                """contenido_celda = re.sub(r'(\b[Ff]ederal \d+),', r'\1 >', contenido_celda)
                contenido_celda = re.sub(r'(\b[Ll]ocal) (\d+)', r'\1 > \2', contenido_celda)"""
                # Buscar "CD" y agregar ">" antes de "CD"
                contenido_celda = re.sub(r'\bMéxico\.\s*(CD\d+)\b', r'México. >\1', contenido_celda)


                # Agregar ", México" al final del texto
                contenido_celda = contenido_celda.replace(" , México", "> , México")

                # Agregar ">" antes de cada A)
                contenido_celda = re.sub(r'(A\))', r'>\1', contenido_celda)
                
                # Eliminar ">" antes de "B)"
                contenido_celda = re.sub(r'>\s*(\nB\))', r'\1', contenido_celda)
                # Eliminar ">" antes de "\nb)"
                contenido_celda = re.sub(r'>\s*(\n b\))', r'\1', contenido_celda)


                # Agregar ">" después de "km", pero no si está seguido por "B)"
                contenido_celda = re.sub(r'(?<!B\))(\bkm\b)', r'\1 >', contenido_celda)

                # Agregar ">" después de "km", pero no si está seguido por "\nB)"
                contenido_celda = re.sub(r'(?<!\n B\))(\bkm\b)', r'\1 >', contenido_celda)

                # Buscar "N Casillas" y agregar ">" antes del número, considerando mayúsculas y minúsculas
                contenido_celda = re.sub(r'(\b\d\s*[Cc]asillas\b)', r'>\1', contenido_celda)

                # Reemplazar al final de la cadena
                for old, new in [("^IZORE:", "ZORE"), ("ZOR:E", "ZORE"), ("^IZORE", "ZORE")]:
                    if contenido_celda.endswith(old):
                        contenido_celda = contenido_celda[:-len(old)] + new
                
                
                # Actualizar el contenido de la celda con los delimitadores y modificaciones agregados
                cell.value = contenido_celda

        # Guardar el libro de trabajo de Excel modificado
        wb.save(filename='excel_final.xlsx')
        
        # Actualizar la interfaz con un mensaje de éxito
        lbl_estado.config(text="Archivo procesado exitosamente.")
    else:
        # Actualizar la interfaz con un mensaje de error
        lbl_estado.config(text="No se seleccionó ningún archivo.")

# Define la función eliminar_urls 
def eliminar_urls(texto):
    return re.sub(r'https?://\S+', '', texto)

# Crear la ventana principal
ventana = tk.Tk()
ventana.title("Cédulas de Mecanismos")


# Crear un widget de etiqueta para mostrar las instrucciones
instrucciones_texto = """\
Instrucciones de uso del programa
1. Seleccionar archivo .txt
2. El archivo excel se guarda en la misma carpeta donde se encuentra el programa, con el nombre "excel_final.xlsx"
3. Es importante guardar el excel después de procesar cada cédula, ya que este se reemplaza automáticamente."""

etiqueta_instrucciones = tk.Label(ventana, text=instrucciones_texto, justify="left")
etiqueta_instrucciones.pack(pady=10)

# Crear un botón para seleccionar el archivo
btn_seleccionar = tk.Button(ventana, text="Seleccionar Archivo", command=procesar_archivo)
btn_seleccionar.pack(pady=10)

# Crear una etiqueta para mostrar el estado del proceso
lbl_estado = tk.Label(ventana, text="")
lbl_estado.pack(pady=5)

# Ejecutar el bucle principal de la ventana
ventana.mainloop()